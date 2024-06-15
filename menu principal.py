import sys
from PyQt5 import QtWidgets
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox
from PyQt5 import QtCore, QtGui, QtWidgets
class UI(QMainWindow):
    def __init__(self):
        super(UI,self).__init__()
        loadUi("appv.ui",self)
        self.show()
        self.actionAjouter_voiture.triggered.connect(lambda:self.ajouter_voiture())
        self.actionSuppression_d_une_voiture_donn_e.triggered.connect(lambda:self.supprimer_mat())
        self.actionSuppression_une_voiture_d_une_marque_donn_e.triggered.connect(lambda:self.supprimer_mar())
        self.actionSuppression_des_voitres_age_10ans.triggered.connect(lambda:self.supprimer_10ans())
        self.actionPrix.triggered.connect(lambda:self.mdprix())
        self.actionetat.triggered.connect(lambda:self.mdetat())
        self.actionCouleur.triggered.connect(lambda:self.mdcouleur())
        self.actionContenu_du_dictionaire_voitures.triggered.connect(lambda:self.contenu_dic_v())
        self.actionRecherche_par_matricule.triggered.connect(lambda:self.rchmat())
        self.actionRecherche_par_marque.triggered.connect(lambda:self.rchmar())
        self.actionRecherche_par_couleur.triggered.connect(lambda:self.rchcol())
        self.actionRecherche_des_voitures_disponibles.triggered.connect(lambda:self.disponible())
        self.actionRecherche_des_voitures_lou_es.triggered.connect(lambda:self.louee())
        self.actionAjouter_un_nouveau_client.triggered.connect(lambda:self.ajouter_client())
        self.actionSupprimer_un_client.triggered.connect(lambda:self.supprimer_c())
        self.actionadresse.triggered.connect(lambda:self.mdadresse())
        self.actiontelephone.triggered.connect(lambda:self.mdtel())
        self.actionmail.triggered.connect(lambda:self.mdmail())
        self.actioncontenu_du_dictionnaire_clients.triggered.connect(lambda:self.contenu_dic_c())
        self.actionRecherhe_par_CIN.triggered.connect(lambda:self.rchcin())
        self.actionenregistrement_fichier_VOITURES.triggered.connect(lambda:self.enregistrement_v())
        self.actionEnregistrement_fichier_CLIENTS.triggered.connect(lambda:self.enregistrement_c())
        self.actionEnregistrement_fichier_LOCATIONS.triggered.connect(lambda:self.enregistrement_l())
        self.actionAjouter_une_nouvelle_locaztion.triggered.connect(lambda:self.ajouter_loc())
        self.actionSupprimer_une_location.triggered.connect(lambda:self.supprimer_l())
        self.actionContenu_du_dictionnaire_locations.triggered.connect(lambda:self.contenu_dic_l())
        self.actionDate_de_location.triggered.connect(lambda:self.mddate())
        self.actionDur_e.triggered.connect(lambda:self.mddure())
        self.actionRecherche_par_CIN.triggered.connect(lambda:self.rchcinloc())
        self.actionRecherche_par_matricule_2.triggered.connect(lambda:self.rchmatloc())
        self.actionRecherche_par_date_donn_e.triggered.connect(lambda:self.rchdate())
        self.actionRecherche_par_dur_e_de_location.triggered.connect(lambda:self.rchdure())
      
    def ajouter_voiture(self):
        from subprocess import call
        call(["python", "idea.py"])       
    def contenu_dic_v(self):
        from subprocess import call
        call(["python", "okay.py"])
    def supprimer_mat(self):
        from subprocess import call
        call(["python", "suppmat.py"])
        
    def supprimer_mar(self):
        from subprocess import call
        call(["python", "supmarque.py"])
        
    def supprimer_10ans(self):
        filename1="voitures.txt"
        l1=[]
        l2=[]
        d={}
        with open(filename1) as f1:
            for line in f1:
                l1=line.split()
                b=l1[0]
                l2=l1[1:]
                d[b]=l2
        s=[]
        for cle,valeur in d.items():
            t=valeur[3]
            x=t[0:4]
            z=int(x)
            if (2022-z >10):
                s.append(cle)
        if(len(s)==0):
            msg=QMessageBox()
            msg.setWindowTitle("Erreur")
            msg.setText("voitures introuvables")
            x=msg.exec_()
        else:
            for w in s:
                del d[w]
            msg=QMessageBox()
            msg.setWindowTitle("Succés")
            msg.setText("voitures supprimées avec succés")
            x=msg.exec_()
            with open ('voitures.txt','w') as f:
                for cle,valeur in d.items():
                    f.write(cle+" ")
                    f.write(valeur[0]+" ")
                    f.write(valeur[1]+" ")
                    f.write(valeur[2]+" ")
                    f.write(valeur[3]+" ")
                    f.write(valeur[4]+'\n')
    def mdprix(self):
        from subprocess import call
        call(["python", "prix.py"])
    def mdetat(self):
        from subprocess import call
        call(["python", "etat.py"])   
    def mdcouleur(self):
        from subprocess import call
        call(["python", "couleur.py"])
    def disponible(self):
        from subprocess import call
        call(["python", "dispo.py"])
    def louee(self):
        from subprocess import call
        call(["python", "loue.py"])
    def rchmat(self):
        from subprocess import call
        call(["python", "rechmatr.py"])
    def rchmar(self):
        from subprocess import call
        call(["python", "rechmar.py"])
    def rchcol(self):
        from subprocess import call
        call(["python", "rechcol.py"])
    def ajouter_client(self):
        from subprocess import call
        call(["python", "idea2.py"])       
    def contenu_dic_c(self):
        from subprocess import call
        call(["python", "okay2.py"])
    def supprimer_c(self):
        from subprocess import call
        call(["python", "suppcl.py"])
    def rchcin(self):
        from subprocess import call
        call(["python", "rechcin.py"])
    def mdadresse(self):
        from subprocess import call
        call(["python", "adresse.py"])  
    def mdtel(self):
        from subprocess import call
        call(["python", "tel.py"])
    def mdmail(self):
        from subprocess import call
        call(["python", "mail.py"])
    
    def enregistrement_v(self):
        filename1="voitures.txt"
        l1=[]
        l2=[]
        d={}
        with open(filename1) as f1:
            for line in f1:
                l1=line.split()
                b=l1[0]
                l2=l1[1:]
                d[b]=l2
        import csv
        with open ('voitures.csv','w',newline="")as f:
            writer=csv.writer(f)
            for cle,valeur in d.items():
                tupl=(cle,valeur[0],valeur[1],valeur[2],valeur[3],valeur[4])
                writer.writerow(tupl)    
        import openpyxl
        wb=openpyxl.Workbook("")
        sheet=wb.active
        sheet['A1'].value='matricule'
        sheet['B1'].value='marque'
        sheet['C1'].value='couleur'
        sheet['D1'].value='etat'
        sheet['E1'].value='date_achat'
        sheet['F1'].value='prix location'
        i=2
        for cle,valeur in d.items():
            a=sheet['A{}'.format(i)]
            b=sheet['B{}'.format(i)]
            c=sheet['C{}'.format(i)]
            d=sheet['D{}'.format(i)]
            e=sheet['E{}'.format(i)]
            f=sheet['F{}'.format(i)]
            a.value=cle
            b.value=valeur[0]
            c.value=valeur[1]
            d.value=valeur[2]
            e.value=valeur[3]
            f.value=valeur[4]
            i=i+1
        wb.save("voitures.xlsx")
        msg=QMessageBox()
        msg.setWindowTitle("Succés")
        msg.setText("fichiers enregistrés avec succés")
        x=msg.exec_()
    
    def enregistrement_c(self):
        filename1="clients.txt"
        l1=[]
        l2=[]
        dc={}
        with open(filename1) as f1:
            for line in f1:
                l1=line.split()
                b=l1[0]
                l2=l1[1:]
                dc[b]=l2
        import csv
        with open ('clients.csv','w',newline="")as f:
            writer=csv.writer(f)
            for cle,valeur in dc.items():
                tupl=(cle,valeur[0],valeur[1],valeur[2],valeur[3],valeur[4],valeur[5])
                writer.writerow(tupl)    
        import openpyxl
        wb=openpyxl.Workbook("")
        sheet=wb.active
        sheet['A1'].value='CIN'
        sheet['B1'].value='nom'
        sheet['C1'].value='prenom'
        sheet['D1'].value='age'
        sheet['E1'].value='adresse'
        sheet['F1'].value='mail'
        sheet['G1'].value='num_tel'
        i=2
        for cle,valeur in dc.items():
            a=sheet['A{}'.format(i)]
            b=sheet['B{}'.format(i)]
            c=sheet['C{}'.format(i)]
            d=sheet['D{}'.format(i)]
            e=sheet['E{}'.format(i)]
            f=sheet['F{}'.format(i)]
            g=sheet['G{}'.format(i)]
            a.value=cle
            b.value=valeur[0]
            c.value=valeur[1]
            d.value=valeur[2]
            e.value=valeur[3]
            f.value=valeur[4]
            g.value=valeur[5]
            i=i+1
        wb.save("clients.xlsx")
        msg=QMessageBox()
        msg.setWindowTitle("Succés")
        msg.setText("fichiers enregistrés avec succés")
        x=msg.exec_()
    def enregistrement_l(self):
        filename1="locations.txt"
        l1=[]
        l2=[]
        dl={}
        with open(filename1) as f1:
            for line in f1:
                l1=line.split()
                b=l1[0]
                c=l1[1]
                a=l1[2]
                l2=l1[3:]
                dl[b,c,a]=l2
        import csv
        with open ('locations.csv','w',newline="")as f:
            writer=csv.writer(f)
            for cle,valeur in dl.items():
                tupl=(cle[0],cle[1],cle[2],valeur[0],valeur[1])
                writer.writerow(tupl) 
        import openpyxl
        wb=openpyxl.Workbook("")
        sheet=wb.active
        sheet['A1'].value='CIN'
        sheet['B1'].value='matricule'
        sheet['C1'].value='date'
        sheet['D1'].value='durée'
        sheet['E1'].value='montant'
        i=2
        for cle,valeur in dl.items():
            a=sheet['A{}'.format(i)]
            b=sheet['B{}'.format(i)]
            c=sheet['C{}'.format(i)]
            d=sheet['D{}'.format(i)]
            e=sheet['E{}'.format(i)]
            a.value=cle[0]
            b.value=cle[1]
            c.value=cle[2]
            d.value=valeur[0]
            e.value=valeur[1]
            i=i+1
        wb.save("locations.xlsx")
        msg=QMessageBox()
        msg.setWindowTitle("Succés")
        msg.setText("fichiers enregistrés avec succés")
        x=msg.exec_()
    def ajouter_loc(self):
        from subprocess import call
        call(["python", "idea3.py"])       
    def contenu_dic_l(self):
        from subprocess import call
        call(["python", "okay3.py"])
    def supprimer_l(self):
        from subprocess import call
        call(["python", "supploc.py"])
        
    def rchcinloc(self):
        from subprocess import call
        call(["python", "rechcinl.py"])
    def rchmatloc(self):
        from subprocess import call
        call(["python", "rechmatl.py"])
    def rchdate(self):
        from subprocess import call
        call(["python", "rechdate.py"])
    def rchdure(self):
        from subprocess import call
        call(["python", "rechdure.py"])
    def mddate(self):
        from subprocess import call
        call(["python", "date.py"])  
    def mddure(self):
        from subprocess import call
        call(["python", "dure.py"])
        
app=QApplication(sys.argv)
UIWindow=UI()
app.exec_()